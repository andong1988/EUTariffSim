"""
数据分析和图表绘制模块
"""

import json
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import logging

# 设置英文字体（避免中文字体问题）
import matplotlib.font_manager as fm

# 禁用字体警告
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib.font_manager")

# 设置字体
def setup_font():
    """设置字体（支持中文和英文）"""
    try:
        # 尝试使用中文字体
        chinese_fonts = [
            'Microsoft YaHei',  # 微软雅黑
            'SimHei',  # 黑体
            'KaiTi',  # 楷体
            'FangSong',  # 仿宋
            'STHeiti',  # 华文黑体
            'STSong',  # 华文宋体
            'PMingLiU',  # 新细明体
            'MingLiU',  # 细明体
        ]
        
        english_fonts = ['Arial', 'DejaVu Sans', 'Liberation Sans', 'sans-serif']
        
        # 优先使用中文字体
        all_fonts = chinese_fonts + english_fonts
        plt.rcParams['font.sans-serif'] = all_fonts
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['axes.unicode_minus'] = False
        
        print(f"Font setup completed. Font list: {all_fonts}")
    except Exception as e:
        print(f"Font setup failed: {e}")
        # 使用默认字体
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

# 设置字体
setup_font()

class SimulationAnalyzer:
    """模拟数据分析器"""
    
    def __init__(self, simulation_data: Dict = None, results_dir: str = "results"):
        self.simulation_data = simulation_data or {}
        self.results_dir = Path(results_dir)
        self.logger = logging.getLogger(__name__)
        
        # 创建图表输出目录
        self.charts_dir = self.results_dir / "charts"
        self.charts_dir.mkdir(exist_ok=True, parents=True)
        
        # 设置图表样式
        try:
            plt.style.use('seaborn-v0_8')
        except:
            # 如果seaborn-v0_8不可用，使用默认样式
            plt.style.use('default')
        try:
            sns.set_palette("husl")
        except:
            pass  # seaborn不可用时忽略
    
    def analyze_and_visualize(self, simulation_report: Dict) -> Dict:
        """分析数据并生成图表"""
        self.logger.info("开始数据分析和图表生成")
        
        analysis_results = {
            "voting_analysis": self._analyze_voting_patterns(simulation_report),
            "theory_analysis": self._analyze_theoretical_patterns(simulation_report),
            "communication_analysis": self._analyze_communication_patterns(simulation_report),
            "accuracy_analysis": self._analyze_prediction_accuracy(simulation_report),
            "weight_analysis": self._analyze_weight_evolution(simulation_report),
            "charts_generated": []
        }
        
        # 生成图表
        charts = self._generate_comprehensive_charts(simulation_report)
        analysis_results["charts_generated"] = charts
        
        # 保存分析结果
        self._save_analysis_results(analysis_results)
        
        self.logger.info(f"数据分析和图表生成完成，共生成{len(charts)}个图表")
        
        return analysis_results
    
    def _analyze_voting_patterns(self, report: Dict) -> Dict:
        """分析投票模式"""
        initial_voting = report["simulation_phases"]["initial_voting"]
        final_voting = report["simulation_phases"]["final_voting"]
        
        initial_votes = initial_voting["votes"]
        final_votes = final_voting["votes"]
        
        # 投票分布分析
        initial_distribution = self._calculate_vote_distribution(initial_votes)
        final_distribution = self._calculate_vote_distribution(final_votes)
        
        # 立场变化分析
        stance_changes = {}
        for country in initial_votes.keys():
            if country in final_votes:
                if initial_votes[country] != final_votes[country]:
                    stance_changes[country] = {
                        "from": initial_votes[country],
                        "to": final_votes[country]
                    }
        
        return {
            "initial_distribution": initial_distribution,
            "final_distribution": final_distribution,
            "stance_changes": stance_changes,
            "change_rate": len(stance_changes) / len(initial_votes),
            "stability_index": 1 - (len(stance_changes) / len(initial_votes))
        }
    
    def _analyze_theoretical_patterns(self, report: Dict) -> Dict:
        """分析理论模式"""
        initial_voting = report["simulation_phases"]["initial_voting"]
        final_voting = report["simulation_phases"]["final_voting"]
        
        theory_patterns = {}
        
        # 分析每个国家的理论得分模式
        for country_id in initial_voting["voting_details"].keys():
            initial_details = initial_voting["voting_details"][country_id]
            final_details = final_voting["voting_details"][country_id]
            
            theory_patterns[country_id] = {
                "initial_theory_scores": initial_details.get("theoretical_factors", {}),
                "final_theory_scores": final_details.get("theoretical_factors", {}),
                "initial_weights": initial_details.get("weights_used", {}),
                "final_weights": final_details.get("weights_used", {}),
                "initial_decision_score": initial_details.get("decision_score", 0),
                "final_decision_score": final_details.get("decision_score", 0)
            }
        
        return theory_patterns
    
    def _analyze_communication_patterns(self, report: Dict) -> Dict:
        """分析沟通模式"""
        communications = report["simulation_phases"]["communication"]
        
        if not communications.get("country_to_country"):
            return {"enabled": False, "reason": "No communication data"}
        
        country_comms = communications["country_to_country"]
        eu_comms = communications.get("eu_commission", [])
        china_comms = communications.get("china", [])
        
        # 沟通网络分析
        communication_network = {}
        for comm in country_comms:
            initiator = comm["from"]
            target = comm["to"]
            
            if initiator not in communication_network:
                communication_network[initiator] = {"outgoing": 0, "incoming": 0, "targets": []}
            
            communication_network[initiator]["outgoing"] += 1
            communication_network[initiator]["targets"].append(target)
            
            if target not in communication_network:
                communication_network[target] = {"outgoing": 0, "incoming": 0, "targets": []}
            
            communication_network[target]["incoming"] += 1
        
        return {
            "enabled": True,
            "country_to_country_count": len(country_comms),
            "eu_commission_count": len(eu_comms),
            "china_count": len(china_comms),
            "total_communications": len(country_comms) + len(eu_comms) + len(china_comms),
            "communication_network": communication_network,
            "retaliation_triggered": communications.get("retaliation", {}).get("triggered", False)
        }
    
    def _analyze_prediction_accuracy(self, report: Dict) -> Dict:
        """分析预测准确率"""
        accuracy_analysis = report["analysis"]["accuracy_analysis"]
        
        country_accuracy = accuracy_analysis.get("country_accuracy", {})
        
        # 按国家分析准确率
        accuracy_by_country = {}
        for country, data in country_accuracy.items():
            accuracy_by_country[country] = {
                "accuracy": data["accuracy"],
                "actual_votes": data["actual_votes"],
                "simulated_votes": data["simulated_votes"],
                "correct_predictions": data["correct_predictions"]
            }
        
        # 整体准确率统计
        overall_accuracy = accuracy_analysis.get("overall_accuracy", 0)
        
        return {
            "overall_accuracy": overall_accuracy,
            "country_accuracy": accuracy_by_country,
            "countries_evaluated": len(country_accuracy),
            "high_accuracy_countries": [c for c, d in country_accuracy.items() if d["accuracy"] >= 0.8],
            "low_accuracy_countries": [c for c, d in country_accuracy.items() if d["accuracy"] < 0.5]
        }
    
    def _analyze_weight_evolution(self, report: Dict) -> Dict:
        """分析权重演化"""
        weight_optimization = report["simulation_phases"]["weight_optimization"]
        
        if not weight_optimization.get("enabled", False):
            return {"enabled": False, "reason": "Weight optimization not enabled"}
        
        optimization_results = weight_optimization.get("results", {})
        
        weight_evolution = {}
        for country, data in optimization_results.items():
            original_weights = data["original_weights"]
            optimized_weights = data["optimized_weights"]
            
            # 计算权重变化
            weight_changes = {}
            for theory in original_weights.keys():
                change = optimized_weights[theory] - original_weights[theory]
                if abs(change) > 0.01:  # 只记录显著变化
                    weight_changes[theory] = change
            
            weight_evolution[country] = {
                "original_weights": original_weights,
                "optimized_weights": optimized_weights,
                "weight_changes": weight_changes,
                "accuracy_improvement": data["improvement"]
            }
        
        return {
            "enabled": True,
            "weight_evolution": weight_evolution,
            "countries_optimized": len(optimization_results),
            "average_improvement": np.mean([d["improvement"] for d in optimization_results.values()]) if optimization_results else 0
        }
    
    def _calculate_vote_distribution(self, votes: Dict) -> Dict:
        """计算投票分布"""
        distribution = {"support": 0, "against": 0, "abstain": 0, "neutral": 0}
        for vote in votes.values():
            distribution[vote] = distribution.get(vote, 0) + 1
        return distribution
    
    def _generate_comprehensive_charts(self, report: Dict) -> List[str]:
        """生成综合图表"""
        charts_generated = []
        
        # 1. 投票结果对比图
        chart1 = self._create_voting_comparison_chart(report)
        if chart1:
            charts_generated.append(chart1)
        
        # 2. 理论权重分布图
        chart2 = self._create_theory_weights_chart(report)
        if chart2:
            charts_generated.append(chart2)
        
        # 3. 理论得分热力图
        chart3 = self._create_theory_scores_heatmap(report)
        if chart3:
            charts_generated.append(chart3)
        
        # 4. 沟通网络图
        chart4 = self._create_communication_network_chart(report)
        if chart4:
            charts_generated.append(chart4)
        
        # 5. 预测准确率图
        chart5 = self._create_accuracy_chart(report)
        if chart5:
            charts_generated.append(chart5)
        
        # 6. 权重演化图
        chart6 = self._create_weight_evolution_chart(report)
        if chart6:
            charts_generated.append(chart6)
        
        # 7. 立场变化图
        chart7 = self._create_stance_change_chart(report)
        if chart7:
            charts_generated.append(chart7)
        
        # 8. 综合分析仪表板
        chart8 = self._create_dashboard(report)
        if chart8:
            charts_generated.append(chart8)
        
        # 9. 沟通内容分析表格
        chart9 = self._create_communication_matrix_chart(report)
        if chart9:
            charts_generated.append(chart9)
        
        return charts_generated
    
    def _create_voting_comparison_chart(self, report: Dict) -> Optional[str]:
        """创建投票结果对比图"""
        try:
            initial_voting = report["simulation_phases"]["initial_voting"]
            final_voting = report["simulation_phases"]["final_voting"]
            
            initial_votes = initial_voting["votes"]
            final_votes = final_voting["votes"]
            
            countries = list(initial_votes.keys())
            initial_positions = [1 if v == "support" else (-1 if v == "against" else 0) for v in initial_votes.values()]
            final_positions = [1 if v == "support" else (-1 if v == "against" else 0) for v in final_votes.values()]
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 8))
            
            # 初始投票分布
            initial_dist = self._calculate_vote_distribution(initial_votes)
            ax1.pie(initial_dist.values(), labels=[f"{k}\n({v} votes)" for k, v in initial_dist.items()], 
                   autopct='%1.1f%%', startangle=90)
            ax1.set_title('Initial Voting Distribution', fontsize=14, fontweight='bold')
            
            # 最终投票分布
            final_dist = self._calculate_vote_distribution(final_votes)
            ax2.pie(final_dist.values(), labels=[f"{k}\n({v} votes)" for k, v in final_dist.items()], 
                   autopct='%1.1f%%', startangle=90)
            ax2.set_title('Final Voting Distribution', fontsize=14, fontweight='bold')
            
            plt.suptitle('EU Tariff Voting Results Comparison', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            filename = f"voting_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建投票对比图失败: {e}")
            return None
    
    def _create_theory_weights_chart(self, report: Dict) -> Optional[str]:
        """创建理论权重分布图"""
        try:
            initial_voting = report["simulation_phases"]["initial_voting"]
            final_voting = report["simulation_phases"]["final_voting"]
            
            # 收集所有理论权重
            theories = ["rational_choice", "two_level_game", "constructivism", "dependency_weaponization"]
            countries = list(initial_voting["voting_details"].keys())
            
            initial_weights_data = {theory: [] for theory in theories}
            final_weights_data = {theory: [] for theory in theories}
            
            for country in countries:
                initial_weights = initial_voting["voting_details"][country].get("weights_used", {})
                final_weights = final_voting["voting_details"][country].get("weights_used", {})
                
                for theory in theories:
                    initial_weights_data[theory].append(initial_weights.get(theory, 0))
                    final_weights_data[theory].append(final_weights.get(theory, 0))
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
            
            # 初始权重分布
            x = np.arange(len(countries))
            width = 0.2
            
            for i, theory in enumerate(theories):
                ax1.bar(x + i*width, initial_weights_data[theory], width, 
                       label=self._get_theory_english_name(theory), alpha=0.8)
            
            ax1.set_xlabel('Country')
            ax1.set_ylabel('Weight')
            ax1.set_title('Initial Theory Weights Distribution')
            ax1.set_xticks(x + width * 1.5)
            ax1.set_xticklabels(countries, rotation=45)
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            
            # 最终权重分布
            for i, theory in enumerate(theories):
                ax2.bar(x + i*width, final_weights_data[theory], width, 
                       label=self._get_theory_english_name(theory), alpha=0.8)
            
            ax2.set_xlabel('Country')
            ax2.set_ylabel('Weight')
            ax2.set_title('Final Theory Weights Distribution')
            ax2.set_xticks(x + width * 1.5)
            ax2.set_xticklabels(countries, rotation=45)
            ax2.legend()
            ax2.grid(True, alpha=0.3)
            
            plt.suptitle('Theory Weights Distribution Comparison', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            filename = f"theory_weights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建理论权重图失败: {e}")
            return None
    
    def _create_theory_scores_heatmap(self, report: Dict) -> Optional[str]:
        """创建理论得分热力图"""
        try:
            initial_voting = report["simulation_phases"]["initial_voting"]
            final_voting = report["simulation_phases"]["final_voting"]
            
            countries = list(initial_voting["voting_details"].keys())
            theories = ["rational_choice", "two_level_games", "constructivism", "weaponized_interdependence"]
            
            # 准备数据
            initial_scores = []
            final_scores = []
            
            for country in countries:
                initial_factors = initial_voting["voting_details"][country].get("theoretical_factors", {})
                final_factors = final_voting["voting_details"][country].get("theoretical_factors", {})
                
                initial_row = [initial_factors.get(theory, 0) for theory in theories]
                final_row = [final_factors.get(theory, 0) for theory in theories]
                
                initial_scores.append(initial_row)
                final_scores.append(final_row)
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 10))
            
            # 初始理论得分热力图
            sns.heatmap(np.array(initial_scores).T, 
                       xticklabels=countries,
                       yticklabels=[self._get_theory_english_name(t) for t in theories],
                       annot=True, fmt='.2f', cmap='RdYlBu_r', ax=ax1,
                       cbar_kws={'label': 'Theory Score'})
            ax1.set_title('Initial Theory Scores Heatmap', fontsize=14, fontweight='bold')
            ax1.set_xlabel('Country')
            ax1.set_ylabel('Theory')
            
            # 最终理论得分热力图
            sns.heatmap(np.array(final_scores).T, 
                       xticklabels=countries,
                       yticklabels=[self._get_theory_english_name(t) for t in theories],
                       annot=True, fmt='.2f', cmap='RdYlBu_r', ax=ax2,
                       cbar_kws={'label': 'Theory Score'})
            ax2.set_title('Final Theory Scores Heatmap', fontsize=14, fontweight='bold')
            ax2.set_xlabel('Country')
            ax2.set_ylabel('Theory')
            
            plt.suptitle('Theory Scores Heatmap Comparison', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            filename = f"theory_scores_heatmap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建理论得分热力图失败: {e}")
            return None
    
    def _detect_communication_intent(self, communication: Dict[str, Any]) -> str:
        """
        识别沟通内容的倾向性（使用细致的关键词匹配）
        
        Args:
            communication: 沟通数据字典
            
        Returns:
            'support' (希望赞同), 'against' (希望反对), 'abstain' (希望弃权)
        """
        # 提取沟通内容
        content = ""
        if "message" in communication:
            msg = communication["message"]
            if isinstance(msg, dict):
                content = msg.get("content", "")
            else:
                content = str(msg)
        elif "content" in communication:
            content = str(communication["content"])
        
        if not content:
            # 如果没有内容，尝试从desired_vote字段获取
            desired_vote = communication.get("desired_vote", "")
            if desired_vote:
                if desired_vote == "support":
                    return "support"
                elif desired_vote == "against":
                    return "against"
                elif desired_vote == "abstain":
                    return "abstain"
            return "support"  # 默认倾向
        
        # 转换为小写进行匹配
        content_lower = content.lower()
        
        # 支持赞同的关键词（细致分类）
        support_keywords = [
            # 中文支持词
            '支持', '赞同', '赞成', '同意', '通过', '批准', '认可', '接受', '采纳',
            '应该', '必须', '必要', '需要', '要求', '关键', '重要', '核心', '首要',
            '有益', '有利', '正面', '积极', '正面', '良好', '优势', '有利条件',
            '建议', '推荐', '提议', '提倡', '主张', '呼吁', '请求',
            '支持', 'backing', 'endorsing', 'endorsed', 'backed',
            # 英文支持词
            'support', 'supportive', 'supporting', 'back', 'backing', 'endorsing', 'endorse', 'endorsed',
            'agree', 'agreement', 'agreed', 'agrees', 'approve', 'approval', 'approved', 'approves',
            'favor', 'favorable', 'favored', 'favoring', 'preferred', 'prefer',
            'must', 'should', 'should', 'necessary', 'necessity', 'require', 'required', 'required',
            'essential', 'critical', 'vital', 'important', 'key', 'crucial', 'fundamental',
            'beneficial', 'benefit', 'benefits', 'advantage', 'advantageous', 'positive', 'positive',
            'help', 'helping', 'helped', 'helps', 'assist', 'assistance', 'assisting', 'assists',
            'aid', 'aiding', 'aided', 'advocate', 'advocating', 'advocates', 'advocated',
            'recommend', 'recommendation', 'recommended', 'recommends', 'propose', 'proposal', 'proposed', 'proposes',
            'suggest', 'suggestion', 'suggested', 'suggests', 'encourage', 'encouragement', 'encouraging', 'encourages',
            'promote', 'promotion', 'promoting', 'promotes', 'champion', 'championing', 'champions'
        ]
        
        # 反对的关键词（细致分类）
        against_keywords = [
            # 中文反对词
            '反对', '不', '拒绝', '否决', '否决权', '阻止', '抵制', '抗议', '抗议',
            '取消', '撤销', '撤回', '终止', '中止', '暂停', '废除', '作废',
            '消极', '负面', '有害', '损害', '伤害', '破坏', '破坏性', '不利',
            '反对票', '投反对', '投反对票',
            # 英文反对词
            'against', 'oppose', 'opposition', 'opposed', 'opposes', 'opposing',
            'reject', 'rejection', 'rejected', 'rejects', 'rebut', 'rebuttal', 'rebuts',
            'disagree', 'disagreement', 'disagreed', 'disagrees', 'dissent', 'dissention',
            'veto', 'vetoing', 'vetoes', 'vetoed', 'block', 'blocking', 'blocked', 'blocks',
            'prevent', 'prevention', 'preventing', 'prevents', 'stop', 'stopping', 'stopped', 'stops',
            'resist', 'resistance', 'resisting', 'resists', 'protest', 'protesting', 'protests',
            'withdraw', 'withdrawal', 'withdrawing', 'withdraws', 'revoke', 'revocation', 'revoking',
            'cancel', 'cancellation', 'cancelling', 'cancelled', 'cancel', 'terminate', 'termination',
            'negative', 'negativity', 'harmful', 'harming', 'harms', 'hurt', 'hurting', 'hurts',
            'damage', 'damaging', 'damaged', 'damages', 'detrimental', 'adverse', 'adversely'
        ]
        
        # 弃权的关键词（细致分类）
        abstain_keywords = [
            # 中文弃权词
            '弃权', '中立', '中立性', '中立立场', '中立态度',
            '观望', '观察', '等待', '等待中', '等待时机',
            '推迟', '延期', '暂缓', '缓期', '延期处理', '暂不处理',
            '犹豫', '犹豫不决', '迟疑', '迟疑不决', '举棋不定',
            '待定', '待定中', '未定', '未确定', '尚未决定',
            '不确定', '不确定', '不明确', '不清楚', '未知',
            '考虑', '考虑中', '斟酌', '权衡', '思考', '思考中',
            '审议', '审查', '研究', '研究', '分析', '分析中',
            '评估', '评估', '评价', '考察', '考察中',
            # 英文弃权词
            'abstain', 'abstention', 'abstaining', 'abstained', 'abstains',
            'neutral', 'neutrality', 'neutralize', 'neutralizing', 'neutralized',
            'wait', 'waiting', 'waits', 'await', 'awaiting', 'awaits',
            'delay', 'delaying', 'delayed', 'delays', 'postpone', 'postponement', 'postponing',
            'defer', 'deferral', 'deferring', 'defers', 'suspend', 'suspension', 'suspending',
            'hesitate', 'hesitation', 'hesitating', 'hesitates', 'undecided', 'undecided',
            'pending', 'pending', 'uncertain', 'uncertainty', 'uncertainly',
            'consider', 'consideration', 'considering', 'considers', 'review', 'reviewing',
            'examine', 'examination', 'examining', 'examine', 'study', 'studying',
            'analyze', 'analysis', 'analyzing', 'analyze', 'evaluate', 'evaluation', 'evaluating'
        ]
        
        # 统计关键词出现次数
        support_count = sum(1 for keyword in support_keywords if keyword in content_lower)
        against_count = sum(1 for keyword in against_keywords if keyword in content_lower)
        abstain_count = sum(1 for keyword in abstain_keywords if keyword in content_lower)
        
        # 根据关键词数量判断倾向性
        if support_count > against_count and support_count > abstain_count:
            return "support"
        elif against_count > support_count and against_count > abstain_count:
            return "against"
        elif abstain_count > support_count and abstain_count > against_count:
            return "abstain"
        elif support_count == against_count and support_count == 0:
            # 如果没有关键词，返回默认倾向support
            return "support"
        elif support_count == against_count and support_count == abstain_count:
            # 如果三个类型关键词数量相同，优先返回support
            return "support"
        else:
            # 如果有两个类型关键词数量相同且最多，优先返回数量多的那个
            max_count = max(support_count, against_count, abstain_count)
            if support_count == max_count:
                return "support"
            elif against_count == max_count:
                return "against"
            else:
                return "abstain"
    
    def _get_intent_color(self, intent: str) -> str:
        """
        根据倾向性获取对应的颜色
        
        Args:
            intent: 倾向性 ('support', 'against', 'abstain', 'neutral')
            
        Returns:
            颜色代码
        """
        color_map = {
            'support': '#90EE90',      # 淡绿色 - 希望赞同
            'against': '#FFB6C1',      # 淡红色 - 希望反对
            'abstain': '#FFFACD',      # 淡黄色 - 希望弃权
            'neutral': '#D3D3D3'       # 灰色 - 中立
        }
        return color_map.get(intent, '#D3D3D3')
    
    def _create_communication_network_chart(self, report: Dict) -> Optional[str]:
        """创建沟通网络图（增强版：包含期望投票类型标注和统计）"""
        try:
            communications = report["simulation_phases"]["communication"]
            
            if not communications.get("country_to_country"):
                return None
            
            country_comms = communications["country_to_country"]
            
            # 构建网络数据
            countries = set()
            for comm in country_comms:
                countries.add(comm["from"])
                countries.add(comm["to"])
            
            countries = list(countries)
            countries.sort()
            
            # 创建邻接矩阵和倾向性矩阵
            n = len(countries)
            adj_matrix = np.zeros((n, n))
            intent_matrix = {}
            
            # 统计每个国家的期望投票类型
            country_intent_stats = {country: {'support': 0, 'against': 0, 'abstain': 0} for country in countries}
            
            for comm in country_comms:
                from_country = comm["from"]
                to_country = comm["to"]
                from_idx = countries.index(from_country)
                to_idx = countries.index(to_country)
                adj_matrix[from_idx][to_idx] = 1
                
                # 识别沟通倾向性
                intent = self._detect_communication_intent(comm)
                intent_matrix[(from_idx, to_idx)] = intent
                
                # 统计期望投票类型
                country_intent_stats[from_country][intent] += 1
            
            fig, ax = plt.subplots(1, 1, figsize=(16, 14))
            
            # 绘制网络图 - 扩大布局范围以容纳标签
            pos = {}
            for i, country in enumerate(countries):
                angle = 2 * np.pi * i / n
                # 增大半径到1.4，为下方标签和边标签留出空间
                pos[country] = (1.4 * np.cos(angle), 1.4 * np.sin(angle))
            
            # 绘制边 - 使用倾向性颜色和标签
            for comm in country_comms:
                from_country = comm["from"]
                to_country = comm["to"]
                from_idx = countries.index(from_country)
                to_idx = countries.index(to_country)
                intent = intent_matrix.get((from_idx, to_idx), 'neutral')
                color = self._get_intent_color(intent)
                
                from_pos = pos[from_country]
                to_pos = pos[to_country]
                
                # 计算箭头的起点和终点（避免箭头与节点重叠）
                dx = to_pos[0] - from_pos[0]
                dy = to_pos[1] - from_pos[1]
                length = np.sqrt(dx**2 + dy**2)
                
                # 箭头从节点边缘开始，在节点边缘结束
                offset = 0.2  # 节点半径的偏移量
                start_x = from_pos[0] + (dx / length) * offset
                start_y = from_pos[1] + (dy / length) * offset
                end_x = to_pos[0] - (dx / length) * offset
                end_y = to_pos[1] - (dy / length) * offset
                
                # 绘制箭头
                ax.arrow(start_x, start_y, 
                       end_x - start_x, end_y - start_y,
                       head_width=0.05, head_length=0.05, 
                       fc=color, ec=color, alpha=0.8, linewidth=2.5, zorder=2)
            
            # 绘制节点
            for country, (x, y) in pos.items():
                ax.scatter(x, y, s=500, c='white', edgecolors='navy', linewidth=3, zorder=5)
                
                # 国家名称放在节点外部，根据y坐标决定在上方还是下方
                label_offset = 0.25
                if y > 0:
                    # 位于上方的圆圈，标签放在上面
                    label_y = y + label_offset
                    va = 'bottom'
                else:
                    # 位于下方的圆圈，标签放在下面
                    label_y = y - label_offset
                    va = 'top'
                
                ax.text(x, label_y, country, ha='center', va=va, fontsize=12, 
                       fontweight='bold', zorder=6, color='navy')
            
            ax.set_xlim(-1.8, 1.8)
            ax.set_ylim(-1.8, 1.8)
            ax.set_aspect('equal')
            ax.set_title('Communication Network with Desired Vote Types\n(S=Support, A=Against, B=Abstain)', 
                        fontsize=18, fontweight='bold', pad=20)
            ax.axis('off')
            
            # 添加倾向性图例
            legend_elements = [
                plt.Rectangle((0, 0), 1, 1, facecolor='#90EE90', edgecolor='navy', 
                             label='Support', alpha=0.8),
                plt.Rectangle((0, 0), 1, 1, facecolor='#FFB6C1', edgecolor='navy', 
                             label='Against', alpha=0.8),
                plt.Rectangle((0, 0), 1, 1, facecolor='#FFFACD', edgecolor='navy', 
                             label='Abstain', alpha=0.8)
            ]
            ax.legend(handles=legend_elements, loc='upper right', 
                    bbox_to_anchor=(1.15, 1), fontsize=13)
            
            # 添加统计信息
            eu_comms = len(communications.get("eu_commission", []))
            china_comms = len(communications.get("china", []))
            retaliation = communications.get("retaliation", {}).get("triggered", False)
            
            # 统计各倾向性的沟通数量
            intent_counts = {'support': 0, 'against': 0, 'abstain': 0}
            for intent in intent_matrix.values():
                intent_counts[intent] = intent_counts.get(intent, 0) + 1
            
            total = sum(intent_counts.values())
            
            stats_text = f"""Statistics:
EU Commission: {eu_comms}
China: {china_comms}
Retaliation: {'Triggered' if retaliation else 'Not Triggered'}

Communication Intents:
Support: {intent_counts['support']} ({intent_counts['support']/total*100:.1f}%)
Against: {intent_counts['against']} ({intent_counts['against']/total*100:.1f}%)
Abstain: {intent_counts['abstain']} ({intent_counts['abstain']/total*100:.1f}%)
Total: {total}"""
            
            ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, 
                   verticalalignment='top', 
                   bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.9, edgecolor='navy'),
                   fontsize=11, family='monospace')
            
            plt.tight_layout()
            
            filename = f"communication_network_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            # 导出CSV文件：网络边详细信息和节点统计
            self._export_network_data(countries, pos, intent_matrix, country_intent_stats, country_comms)
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建沟通网络图失败: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None
    
    def _create_accuracy_chart(self, report: Dict) -> Optional[str]:
        """创建预测准确率图"""
        try:
            accuracy_analysis = report["analysis"]["accuracy_analysis"]
            country_accuracy = accuracy_analysis.get("country_accuracy", {})
            
            if not country_accuracy:
                return None
            
            countries = list(country_accuracy.keys())
            accuracies = [country_accuracy[c]["accuracy"] for c in countries]
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
            
            # 准确率柱状图
            bars = ax1.bar(countries, accuracies, color=['#2ecc71' if acc >= 0.8 else '#f39c12' if acc >= 0.5 else '#e74c3c' for acc in accuracies])
            ax1.set_xlabel('Country')
            ax1.set_ylabel('Accuracy')
            ax1.set_title('Prediction Accuracy by Country', fontsize=14, fontweight='bold')
            ax1.set_ylim(0, 1)
            ax1.grid(True, alpha=0.3)
            
            # 添加数值标签
            for bar, acc in zip(bars, accuracies):
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 0.02,
                        f'{acc:.2f}', ha='center', va='bottom')
            
            # 准确率分布
            overall_accuracy = accuracy_analysis.get("overall_accuracy", 0)
            ax2.hist(accuracies, bins=5, alpha=0.7, color='skyblue', edgecolor='black')
            ax2.axvline(overall_accuracy, color='red', linestyle='--', linewidth=2, label=f'Overall Accuracy: {overall_accuracy:.2f}')
            ax2.set_xlabel('Accuracy')
            ax2.set_ylabel('Number of Countries')
            ax2.set_title('Accuracy Distribution', fontsize=14, fontweight='bold')
            ax2.legend()
            ax2.grid(True, alpha=0.3)
            
            plt.suptitle('Prediction Accuracy Analysis', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            filename = f"accuracy_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建准确率图失败: {e}")
            return None
    
    def _create_weight_evolution_chart(self, report: Dict) -> Optional[str]:
        """创建权重演化图"""
        try:
            weight_optimization = report["simulation_phases"]["weight_optimization"]
            
            if not weight_optimization.get("enabled", False):
                return None
            
            optimization_results = weight_optimization.get("results", {})
            
            if not optimization_results:
                return None
            
            countries = list(optimization_results.keys())
            theories = ["rational_choice", "two_level_game", "constructivism", "dependency_weaponization"]
            
            fig, axes = plt.subplots(2, 2, figsize=(16, 12))
            axes = axes.flatten()
            
            for i, theory in enumerate(theories):
                ax = axes[i]
                
                original_weights = [optimization_results[c]["original_weights"].get(theory, 0) for c in countries]
                optimized_weights = [optimization_results[c]["optimized_weights"].get(theory, 0) for c in countries]
                
                x = np.arange(len(countries))
                width = 0.35
                
                ax.bar(x - width/2, original_weights, width, label='Before Optimization', alpha=0.8)
                ax.bar(x + width/2, optimized_weights, width, label='After Optimization', alpha=0.8)
                
                ax.set_xlabel('Country')
                ax.set_ylabel('Weight')
                ax.set_title(f'{self._get_theory_english_name(theory)} Weight Changes')
                ax.set_xticks(x)
                ax.set_xticklabels(countries, rotation=45)
                ax.legend()
                ax.grid(True, alpha=0.3)
            
            plt.suptitle('Theory Weight Optimization Comparison', fontsize=16, fontweight='bold')
            plt.tight_layout()
            
            filename = f"weight_evolution_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建权重演化图失败: {e}")
            return None
    
    def _create_stance_change_chart(self, report: Dict) -> Optional[str]:
        """创建立场变化图"""
        try:
            initial_voting = report["simulation_phases"]["initial_voting"]
            final_voting = report["simulation_phases"]["final_voting"]
            
            initial_votes = initial_voting["votes"]
            final_votes = final_voting["votes"]
            
            countries = list(initial_votes.keys())
            
            # 立场编码
            stance_map = {"support": 1, "against": -1, "abstain": 0, "neutral": 0}
            
            initial_positions = [stance_map.get(initial_votes[c], 0) for c in countries]
            final_positions = [stance_map.get(final_votes[c], 0) for c in countries]
            
            fig, ax = plt.subplots(1, 1, figsize=(12, 8))
            
            # 绘制立场变化
            x = np.arange(len(countries))
            
            for i, country in enumerate(countries):
                initial_pos = initial_positions[i]
                final_pos = final_positions[i]
                
                if initial_pos != final_pos:
                    # 绘制变化箭头
                    ax.arrow(i, initial_pos, 0, final_pos - initial_pos, 
                            head_width=0.1, head_length=0.1, fc='red', ec='red', alpha=0.7)
                
                # 标记位置
                ax.scatter(i, initial_pos, s=100, c='blue', marker='o', label='Initial Stance' if i == 0 else "")
                ax.scatter(i, final_pos, s=100, c='green', marker='s', label='Final Stance' if i == 0 else "")
            
            ax.set_xlabel('Country')
            ax.set_ylabel('Stance')
            ax.set_title('Country Stance Changes', fontsize=16, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(countries, rotation=45)
            ax.set_yticks([-1, 0, 1])
            ax.set_yticklabels(['Against', 'Abstain/Neutral', 'Support'])
            ax.grid(True, alpha=0.3)
            ax.legend()
            
            # 添加统计信息
            stance_changes = sum(1 for i in range(len(countries)) if initial_positions[i] != final_positions[i])
            change_rate = stance_changes / len(countries)
            
            stats_text = f"Stance Changes: {stance_changes}/{len(countries)} ({change_rate:.1%})\nStability Index: {1-change_rate:.2f}"
            ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, 
                   verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            
            plt.tight_layout()
            
            filename = f"stance_change_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建立场变化图失败: {e}")
            return None
    
    def _create_communication_matrix_chart(self, report: Dict) -> Optional[str]:
        """创建沟通内容分析表格（增强版：包含期望投票类型标注和统计分析）"""
        try:
            communications = report["simulation_phases"]["communication"]
            
            if not communications.get("country_to_country"):
                return None
            
            country_comms = communications["country_to_country"]
            
            # 获取所有国家
            countries = set()
            for comm in country_comms:
                countries.add(comm["from"])
                countries.add(comm["to"])
            
            countries = sorted(list(countries))
            
            # 初始化沟通矩阵（发送国家×接收国家）
            n = len(countries)
            comm_matrix = np.zeros((n, n), dtype=int)
            
            # 初始化加权得分矩阵
            weighted_score_matrix = np.zeros((n, n), dtype=float)
            
            # 初始化期望投票类型矩阵
            intent_counts_matrix = {}  # {(from_idx, to_idx): {'support': 0, 'against': 0, 'abstain': 0}}
            
            # 计数器和得分累加器
            count_matrix = np.zeros((n, n), dtype=int)
            score_sum_matrix = np.zeros((n, n), dtype=float)
            
            # 填充矩阵
            for comm in country_comms:
                from_country = comm["from"]
                to_country = comm["to"]
                
                from_idx = countries.index(from_country)
                to_idx = countries.index(to_country)
                
                # 统计沟通次数
                count_matrix[from_idx][to_idx] += 1
                
                # 识别沟通倾向性
                intent = self._detect_communication_intent(comm)
                
                # 统计期望投票类型
                key = (from_idx, to_idx)
                if key not in intent_counts_matrix:
                    intent_counts_matrix[key] = {'support': 0, 'against': 0, 'abstain': 0}
                intent_counts_matrix[key][intent] += 1
                
                # 计算加权得分
                # 赞成=1分，反对=-1分，弃权=0.4分
                if intent == "support":
                    score = 1.0
                elif intent == "against":
                    score = -1.0
                elif intent == "abstain":
                    score = 0.4
                else:
                    score = 0.0
                
                score_sum_matrix[from_idx][to_idx] += score
            
            # 计算平均加权得分
            for i in range(n):
                for j in range(n):
                    if count_matrix[i][j] > 0:
                        weighted_score_matrix[i][j] = score_sum_matrix[i][j] / count_matrix[i][j]
            
            # 创建颜色矩阵（基于平均加权得分）
            # ≥1分：淡绿色；<0分：淡红色；其他：淡黄色
            color_matrix = np.empty((n, n, 4))  # RGBA
            
            for i in range(n):
                for j in range(n):
                    if count_matrix[i][j] == 0:
                        # 无沟通，使用浅灰色
                        color_matrix[i][j] = [0.9, 0.9, 0.9, 0.5]
                    elif weighted_score_matrix[i][j] >= 1.0:
                        # 淡绿色
                        color_matrix[i][j] = [0.56, 0.93, 0.56, 0.8]  # #90EE90
                    elif weighted_score_matrix[i][j] < 0:
                        # 淡红色
                        color_matrix[i][j] = [1.0, 0.71, 0.76, 0.8]  # #FFB6C1
                    else:
                        # 淡黄色
                        color_matrix[i][j] = [1.0, 0.98, 0.8, 0.8]  # #FFFACD
            
            # 创建表格 - 减小宽度以避免列过宽
            fig, ax = plt.subplots(figsize=(10, 12))
            
            # 隐藏坐标轴
            ax.axis('tight')
            ax.axis('off')
            
            # 创建表格（包含期望投票类型标注）
            table_data = []
            for i in range(n):
                row_data = []
                for j in range(n):
                    count = count_matrix[i][j]
                    if count > 0:
                        score = weighted_score_matrix[i][j]
                        
                        # 获取期望投票类型分布
                        key = (i, j)
                        intents = intent_counts_matrix.get(key, {'support': 0, 'against': 0, 'abstain': 0})
                        
                        # 确定主要倾向性
                        max_intent = max(intents.items(), key=lambda x: x[1])[0]
                        intent_label = {'support': 'S', 'against': 'A', 'abstain': 'B'}.get(max_intent, '')
                        
                        # 构建单元格文本
                        if count == 1:
                            # 只有一次沟通，显示期望投票类型
                            row_data.append(f"{intent_label}\n({score:+.1f})")
                        else:
                            # 多次沟通，显示次数和主要倾向性
                            support_pct = intents['support'] / count * 100
                            against_pct = intents['against'] / count * 100
                            abstain_pct = intents['abstain'] / count * 100
                            row_data.append(f"{count}\n{intent_label}\n({score:+.1f})")
                    else:
                        row_data.append("-")
                table_data.append(row_data)
            
            # 添加行列标签
            table = ax.table(cellText=table_data,
                           rowLabels=countries,
                           colLabels=countries,
                           cellLoc='center',
                           loc='center',
                           cellColours=color_matrix.tolist())
            
            # 设置表格样式 - 调整缩放以减小列宽
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(0.75, 1.3)
            
            # 设置标题行样式
            for col in range(n):
                cell = table[0, col]
                cell.set_facecolor('#4CAF50')
                cell.set_text_props(weight='bold', color='white')
            
            # 设置行首样式
            for row in range(n):
                cell = table[row + 1, -1]  # 行首
                cell.set_facecolor('#4CAF50')
                cell.set_text_props(weight='bold', color='white')
            
            # 设置单元格边框
            for i in range(n + 1):
                for j in range(n):
                    cell = table[i, j]
                    cell.set_edgecolor('#333333')
                    cell.set_linewidth(0.5)
            
            ax.set_title('Communication Matrix with Desired Vote Types\n(S=Support, A=Against, B=Abstain)',
                        fontsize=16, fontweight='bold', pad=20)
            
            # 添加图例说明
            legend_text = """Cell Format:
  Single: Intent (Score)
  Multiple: Count (Score)

Intent Labels:
  S = Support
  A = Against
  B = Abstain

Color Coding:
  Light Green (≥1.0): Strong Support
  Light Red (<0.0): Against
  Light Yellow (0.0-1.0): Abstain/Neutral
  Light Gray: No Communication

Score Calculation:
  Support = +1.0
  Against = -1.0
  Abstain = +0.4"""
            
            fig.text(0.95, 0.5, legend_text, transform=fig.transFigure,
                    fontsize=9, verticalalignment='center',
                    bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            
            plt.tight_layout(rect=[0, 0, 0.92, 1])
            
            filename = f"communication_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            # 同时导出Excel格式和CSV数据
            self._export_communication_matrix_to_excel(count_matrix, weighted_score_matrix, countries, intent_counts_matrix)
            self._export_communication_intent_statistics(countries, intent_counts_matrix, count_matrix, country_comms)
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建沟通矩阵图失败: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None
    
    def _create_dashboard(self, report: Dict) -> Optional[str]:
        """创建综合分析仪表板"""
        try:
            fig = plt.figure(figsize=(20, 16))
            
            # 创建网格布局
            gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
            
            # 1. 整体准确率 (左上)
            ax1 = fig.add_subplot(gs[0, 0])
            overall_accuracy = report["analysis"]["accuracy_analysis"].get("overall_accuracy", 0)
            
            # 创建仪表盘样式
            theta = np.linspace(0, np.pi, 100)
            r_outer = 1
            r_inner = 0.7
            
            # 绘制半圆仪表盘
            ax1.plot(np.cos(theta), np.sin(theta), 'k-', linewidth=2)
            ax1.plot([np.cos(theta[0]), np.cos(theta[-1])], [np.sin(theta[0]), np.sin(theta[-1])], 'k-', linewidth=2)
            
            # 绘制指针
            angle = np.pi * (1 - overall_accuracy)
            ax1.plot([0, 0.8 * np.cos(angle)], [0, 0.8 * np.sin(angle)], 'r-', linewidth=3)
            ax1.scatter(0, 0, s=100, c='black', zorder=5)
            
            # 添加刻度和标签
            for i in range(11):
                angle = np.pi * i / 10
                x = 0.9 * np.cos(angle)
                y = 0.9 * np.sin(angle)
                ax1.plot([0.7 * np.cos(angle), x], [0.7 * np.sin(angle), y], 'k-', linewidth=1)
                if i % 2 == 0:
                    ax1.text(1.1 * np.cos(angle), 1.1 * np.sin(angle), f'{i*10}%', 
                            ha='center', va='center', fontsize=8)
            
            ax1.set_xlim(-1.3, 1.3)
            ax1.set_ylim(-0.3, 1.3)
            ax1.set_aspect('equal')
            ax1.set_title(f'Overall Prediction Accuracy\n{overall_accuracy:.1%}', fontsize=12, fontweight='bold')
            ax1.axis('off')
            
            # 2. 投票分布 (中上)
            ax2 = fig.add_subplot(gs[0, 1])
            initial_voting = report["simulation_phases"]["initial_voting"]
            final_voting = report["simulation_phases"]["final_voting"]
            
            initial_dist = self._calculate_vote_distribution(initial_voting["votes"])
            final_dist = self._calculate_vote_distribution(final_voting["votes"])
            
            labels = list(initial_dist.keys())
            initial_values = list(initial_dist.values())
            final_values = list(final_dist.values())
            
            x = np.arange(len(labels))
            width = 0.35
            
            ax2.bar(x - width/2, initial_values, width, label='Initial Voting', alpha=0.8)
            ax2.bar(x + width/2, final_values, width, label='Final Voting', alpha=0.8)
            
            ax2.set_xlabel('Voting Stance')
            ax2.set_ylabel('Number of Votes')
            ax2.set_title('Voting Distribution Comparison')
            ax2.set_xticks(x)
            ax2.set_xticklabels(labels)
            ax2.legend()
            ax2.grid(True, alpha=0.3)
            
            # 3. 沟通统计 (右上)
            ax3 = fig.add_subplot(gs[0, 2])
            communications = report["simulation_phases"]["communication"]
            
            comm_types = ['Country-to-Country', 'EU Commission', 'China']
            comm_counts = [
                len(communications.get("country_to_country", [])),
                len(communications.get("eu_commission", [])),
                len(communications.get("china", []))
            ]
            
            colors = ['#ff9999', '#66b3ff', '#99ff99']
            ax3.pie(comm_counts, labels=comm_types, autopct='%1.1f%%', colors=colors, startangle=90)
            ax3.set_title('Communication Type Distribution')
            
            # 4. 理论权重平均分布 (左中)
            ax4 = fig.add_subplot(gs[1, 0])
            
            # 计算平均权重
            theories = ["rational_choice", "two_level_game", "constructivism", "dependency_weaponization"]
            avg_weights = []
            
            for country in initial_voting["voting_details"].keys():
                weights = initial_voting["voting_details"][country].get("weights_used", {})
                for i, theory in enumerate(theories):
                    if len(avg_weights) <= i:
                        avg_weights.append(0)
                    avg_weights[i] += weights.get(theory, 0)
            
            avg_weights = [w / len(initial_voting["voting_details"]) for w in avg_weights]
            
            theory_names = [self._get_theory_english_name(t) for t in theories]
            colors = plt.cm.Set3(np.linspace(0, 1, len(theory_names)))
            
            ax4.pie(avg_weights, labels=theory_names, autopct='%1.1f%%', colors=colors, startangle=90)
            ax4.set_title('Average Theory Weights Distribution')
            
            # 5. 立场变化统计 (中中)
            ax5 = fig.add_subplot(gs[1, 1])
            
            stance_changes = 0
            for country in initial_voting["votes"].keys():
                if initial_voting["votes"][country] != final_voting["votes"][country]:
                    stance_changes += 1
            
            stability_rate = (len(initial_voting["votes"]) - stance_changes) / len(initial_voting["votes"])
            
            categories = ['Stable Stance', 'Changed Stance']
            values = [stability_rate, 1 - stability_rate]
            colors = ['#2ecc71', '#e74c3c']
            
            ax5.bar(categories, values, color=colors, alpha=0.8)
            ax5.set_ylabel('Proportion')
            ax5.set_title('Stance Stability')
            ax5.set_ylim(0, 1)
            ax5.grid(True, alpha=0.3)
            
            # 添加数值标签
            for i, v in enumerate(values):
                ax5.text(i, v + 0.02, f'{v:.1%}', ha='center', va='bottom')
            
            # 6. 权重优化效果 (右中)
            ax6 = fig.add_subplot(gs[1, 2])
            weight_optimization = report["simulation_phases"]["weight_optimization"]
            
            if weight_optimization.get("enabled", False):
                optimization_results = weight_optimization.get("results", {})
                improvements = [data["improvement"] for data in optimization_results.values()]
                
                if improvements:
                    ax6.hist(improvements, bins=10, alpha=0.7, color='skyblue', edgecolor='black')
                    ax6.axvline(np.mean(improvements), color='red', linestyle='--', 
                               label=f'Avg Improvement: {np.mean(improvements):.3f}')
                    ax6.set_xlabel('Accuracy Improvement')
                    ax6.set_ylabel('Number of Countries')
                    ax6.set_title('Weight Optimization Effect')
                    ax6.legend()
                    ax6.grid(True, alpha=0.3)
                else:
                    ax6.text(0.5, 0.5, 'No Weight Optimization Data', ha='center', va='center', 
                            transform=ax6.transAxes, fontsize=12)
                    ax6.set_title('Weight Optimization Effect')
            else:
                ax6.text(0.5, 0.5, 'Weight Optimization Not Enabled', ha='center', va='center', 
                        transform=ax6.transAxes, fontsize=12)
                ax6.set_title('Weight Optimization Effect')
            
            # 7. 关键指标总结 (下方横跨)
            ax7 = fig.add_subplot(gs[2, :])
            ax7.axis('off')
            
            # 收集关键指标
            total_comms = (len(communications.get('country_to_country', [])) + 
                          len(communications.get('eu_commission', [])) + 
                          len(communications.get('china', [])))
            metrics = [
                f"Participating Countries: {len(initial_voting['votes'])}",
                f"Overall Accuracy: {overall_accuracy:.1%}",
                f"Stance Stability: {stability_rate:.1%}",
                f"Total Communications: {total_comms}",
                f"Retaliation Measures: {'Triggered' if communications.get('retaliation', {}).get('triggered', False) else 'Not Triggered'}",
                f"Weight Optimization: {'Enabled' if weight_optimization.get('enabled', False) else 'Disabled'}"
            ]
            
            # 创建表格
            table_data = []
            for i, metric in enumerate(metrics):
                if i % 2 == 0:
                    table_data.append([metric, ""])
                else:
                    table_data[-1][1] = metric
            
            table = ax7.table(cellText=table_data, 
                           colLabels=['Key Metrics', 'Key Metrics'],
                           cellLoc='center',
                           loc='center',
                           colWidths=[0.5, 0.5])
            
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            table.scale(1, 2)
            
            # 设置表格样式
            for i in range(len(metrics)):
                row = i // 2
                col = i % 2
                cell = table[(row + 1) if row < len(table_data) else 0, col]
                cell.set_facecolor('#f0f0f0' if i % 2 == 0 else 'white')
                cell.set_edgecolor('black')
            
            # 设置标题行
            for col in range(2):
                cell = table[0, col]
                cell.set_facecolor('#4CAF50')
                cell.set_text_props(weight='bold', color='white')
            
            plt.suptitle('EU Tariff Simulation Comprehensive Analysis Dashboard', fontsize=18, fontweight='bold')
            
            filename = f"dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            filepath = self.charts_dir / filename
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"创建仪表板失败: {e}")
            return None
    
    def _get_theory_english_name(self, theory: str) -> str:
        """获取理论英文名称"""
        theory_names = {
            "rational_choice": "Rational Choice",
            "two_level_game": "Two-Level Game",
            "two_level_games": "Two-Level Game",
            "constructivism": "Constructivism",
            "dependency_weaponization": "Dependency Weaponization",
            "weaponized_interdependence": "Weaponized Interdependence"
        }
        return theory_names.get(theory, theory.replace("_", " ").title())
    
    def _export_communication_matrix_to_excel(self, count_matrix: np.ndarray, 
                                           weighted_score_matrix: np.ndarray, 
                                           countries: List[str],
                                           intent_counts_matrix: Dict):
        """导出沟通矩阵到Excel格式"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"communication_matrix_{timestamp}.xlsx"
            filepath = self.results_dir / filename
            
            # 创建数据目录
            data_dir = self.results_dir / "data"
            data_dir.mkdir(exist_ok=True, parents=True)
            
            # 使用pandas创建DataFrame并导出到Excel
            df_counts = pd.DataFrame(count_matrix, index=countries, columns=countries)
            df_scores = pd.DataFrame(weighted_score_matrix, index=countries, columns=countries)
            
            # 创建Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 写入沟通次数表
                df_counts.to_excel(writer, sheet_name='Communication Counts', engine='openpyxl')
                
                # 写入加权得分表
                df_scores.to_excel(writer, sheet_name='Weighted Scores', engine='openpyxl')
                
                # 写入期望投票类型分析
                self._write_intent_analysis_sheet(writer, countries, intent_counts_matrix, count_matrix)
            
            # 使用openpyxl美化Excel表格
            wb = load_workbook(filepath)
            
            # 定义颜色
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            light_red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            border = Border(
                left=Side(style='thin', color='333333'),
                right=Side(style='thin', color='333333'),
                top=Side(style='thin', color='333333'),
                bottom=Side(style='thin', color='333333')
            )
            
            # 美化沟通次数表
            ws1 = wb['Communication Counts']
            for row in ws1.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 判断单元格位置并应用颜色
                    if cell.row == 1 or cell.column == 1:
                        # 标题行或首列
                        if cell.row == 1 and cell.column == 1:
                            # 左上角空白单元格
                            cell.fill = header_fill
                            cell.font = Font(bold=True, color='FFFFFF')
                        elif cell.row == 1:
                            # 标题行
                            cell.fill = header_fill
                            cell.font = Font(bold=True, color='FFFFFF')
                        else:
                            # 首列（国家名称）
                            cell.font = Font(bold=True)
                    else:
                        # 数据单元格
                        row_idx = cell.row - 2
                        col_idx = cell.column - 2
                        
                        if 0 <= row_idx < len(count_matrix) and 0 <= col_idx < len(count_matrix):
                            count = count_matrix[row_idx][col_idx]
                            if count == 0:
                                cell.fill = gray_fill
                            else:
                                score = weighted_score_matrix[row_idx][col_idx]
                                if score >= 1.0:
                                    cell.fill = light_green_fill
                                elif score < 0:
                                    cell.fill = light_red_fill
                                else:
                                    cell.fill = light_yellow_fill
            
            # 美化加权得分表
            ws2 = wb['Weighted Scores']
            for row in ws2.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if cell.row == 1 or cell.column == 1:
                        if cell.row == 1 and cell.column == 1:
                            cell.fill = header_fill
                            cell.font = Font(bold=True, color='FFFFFF')
                        elif cell.row == 1:
                            cell.fill = header_fill
                            cell.font = Font(bold=True, color='FFFFFF')
                        else:
                            cell.font = Font(bold=True)
                    else:
                        row_idx = cell.row - 2
                        col_idx = cell.column - 2
                        
                        if 0 <= row_idx < len(weighted_score_matrix) and 0 <= col_idx < len(weighted_score_matrix):
                            score = weighted_score_matrix[row_idx][col_idx]
                            count = count_matrix[row_idx][col_idx]
                            
                            if count == 0:
                                cell.fill = gray_fill
                            elif score >= 1.0:
                                cell.fill = light_green_fill
                            elif score < 0:
                                cell.fill = light_red_fill
                            else:
                                cell.fill = light_yellow_fill
            
            # 添加说明sheet
            ws3 = wb.create_sheet('Legend')
            ws3['A1'] = 'Communication Matrix Legend'
            ws3['A1'].font = Font(bold=True, size=14)
            
            ws3['A3'] = 'Cell Format:'
            ws3['A4'] = '- Communication Counts sheet: Shows number of communications'
            ws3['A5'] = '- Weighted Scores sheet: Shows weighted average score'
            
            ws3['A7'] = 'Color Coding (Based on Weighted Score):'
            ws3['A8'] = 'Light Green (≥1.0): Strong Support Tendency'
            ws3['A8'].fill = light_green_fill
            
            ws3['A9'] = 'Light Red (<0.0): Against Tendency'
            ws3['A9'].fill = light_red_fill
            
            ws3['A10'] = 'Light Yellow (0.0-1.0): Neutral/Abstain Tendency'
            ws3['A10'].fill = light_yellow_fill
            
            ws3['A11'] = 'Light Gray: No Communication'
            ws3['A11'].fill = gray_fill
            
            ws3['A13'] = 'Score Calculation:'
            ws3['A14'] = 'Support = +1.0'
            ws3['A15'] = 'Against = -1.0'
            ws3['A16'] = 'Abstain = +0.4'
            ws3['A17'] = 'Neutral = 0.0'
            
            # 调整列宽
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            self.logger.info(f"沟通矩阵已导出到Excel: {filepath}")
            
        except ImportError:
            self.logger.warning("openpyxl not installed, skipping Excel export")
            self.logger.warning("Install with: pip install openpyxl")
        except Exception as e:
            self.logger.error(f"导出Excel失败: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _write_intent_analysis_sheet(self, writer, countries: List[str], 
                                   intent_counts_matrix: Dict, 
                                   count_matrix: np.ndarray):
        """
        写入期望投票类型分析sheet到Excel
        
        Args:
            writer: ExcelWriter对象
            countries: 国家列表
            intent_counts_matrix: 期望投票类型计数矩阵
            count_matrix: 沟通次数矩阵
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 准备数据：每个国家发起的期望投票类型统计
        country_intent_data = []
        for i, country in enumerate(countries):
            total_outgoing = 0
            support_outgoing = 0
            against_outgoing = 0
            abstain_outgoing = 0
            
            total_incoming = 0
            support_incoming = 0
            against_incoming = 0
            abstain_incoming = 0
            
            for j, target_country in enumerate(countries):
                if count_matrix[i][j] > 0:
                    # 发起方统计
                    key = (i, j)
                    intents = intent_counts_matrix.get(key, {'support': 0, 'against': 0, 'abstain': 0})
                    count = count_matrix[i][j]
                    
                    total_outgoing += count
                    support_outgoing += intents['support']
                    against_outgoing += intents['against']
                    abstain_outgoing += intents['abstain']
                
                if count_matrix[j][i] > 0:
                    # 接收方统计
                    key_in = (j, i)
                    intents_in = intent_counts_matrix.get(key_in, {'support': 0, 'against': 0, 'abstain': 0})
                    count_in = count_matrix[j][i]
                    
                    total_incoming += count_in
                    support_incoming += intents_in['support']
                    against_incoming += intents_in['against']
                    abstain_incoming += intents_in['abstain']
            
            country_intent_data.append({
                'Country': country,
                'Total_Outgoing': total_outgoing,
                'Outgoing_Support': support_outgoing,
                'Outgoing_Against': against_outgoing,
                'Outgoing_Abstain': abstain_outgoing,
                'Outgoing_Support_Pct': (support_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0,
                'Outgoing_Against_Pct': (against_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0,
                'Outgoing_Abstain_Pct': (abstain_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0,
                'Total_Incoming': total_incoming,
                'Incoming_Support': support_incoming,
                'Incoming_Against': against_incoming,
                'Incoming_Abstain': abstain_incoming,
                'Incoming_Support_Pct': (support_incoming / total_incoming * 100) if total_incoming > 0 else 0,
                'Incoming_Against_Pct': (against_incoming / total_incoming * 100) if total_incoming > 0 else 0,
                'Incoming_Abstain_Pct': (abstain_incoming / total_incoming * 100) if total_incoming > 0 else 0
            })
        
        # 创建DataFrame并写入Excel
        df_intent = pd.DataFrame(country_intent_data)
        df_intent.to_excel(writer, sheet_name='Intent Analysis by Country', index=False, engine='openpyxl')
        
        self.logger.info("已写入期望投票类型分析sheet")
    
    def _export_communication_intent_statistics(self, countries: List[str], 
                                           intent_counts_matrix: Dict,
                                           count_matrix: np.ndarray,
                                           country_comms: List[Dict]):
        """
        导出沟通意图统计到CSV文件
        
        Args:
            countries: 国家列表
            intent_counts_matrix: 期望投票类型计数矩阵
            count_matrix: 沟通次数矩阵
            country_comms: 沟通记录列表
        """
        # 创建数据目录
        data_dir = self.results_dir / "data"
        data_dir.mkdir(exist_ok=True, parents=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. 每个国家发起的期望投票类型统计CSV
        country_intent_data = []
        for i, country in enumerate(countries):
            total_outgoing = 0
            support_outgoing = 0
            against_outgoing = 0
            abstain_outgoing = 0
            
            for j, target_country in enumerate(countries):
                if count_matrix[i][j] > 0:
                    key = (i, j)
                    intents = intent_counts_matrix.get(key, {'support': 0, 'against': 0, 'abstain': 0})
                    count = count_matrix[i][j]
                    
                    total_outgoing += count
                    support_outgoing += intents['support']
                    against_outgoing += intents['against']
                    abstain_outgoing += intents['abstain']
            
            country_intent_data.append({
                'Country': country,
                'Total_Outgoing': total_outgoing,
                'Outgoing_Support': support_outgoing,
                'Outgoing_Against': against_outgoing,
                'Outgoing_Abstain': abstain_outgoing,
                'Outgoing_Support_Pct': (support_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0,
                'Outgoing_Against_Pct': (against_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0,
                "Outgoing_Abstain_Pct": (abstain_outgoing / total_outgoing * 100) if total_outgoing > 0 else 0
            })
        
        # 导出每个国家的期望投票统计
        csv1_filename = f"communication_intent_by_country_{timestamp}.csv"
        csv1_filepath = data_dir / csv1_filename
        df_country_intent = pd.DataFrame(country_intent_data)
        df_country_intent.to_csv(csv1_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出每个国家的期望投票统计: {csv1_filepath}")
        
        # 2. 双边沟通期望投票类型统计CSV
        bilateral_intent_data = []
        for comm in country_comms:
            from_country = comm["from"]
            to_country = comm["to"]
            
            # 识别沟通意图
            intent = self._detect_communication_intent(comm)
            intent_label = {'support': 'Support', 'against': 'Against', 'abstain': 'Abstain'}.get(intent, 'Unknown')
            
            bilateral_intent_data.append({
                'From_Country': from_country,
                'To_Country': to_country,
                'Intent_Type': intent_label,
                'Intent_Code': intent,
                'Message': comm.get("message", comm.get("content", "")) if isinstance(comm.get("message"), str) else str(comm.get("message", "")),
                'Desired_Vote': comm.get("desired_vote", ""),
                'Timestamp': comm.get("timestamp", "")
            })
        
        # 导出双边沟通期望投票统计
        csv2_filename = f"communication_bilateral_intent_{timestamp}.csv"
        csv2_filepath = data_dir / csv2_filename
        df_bilateral_intent = pd.DataFrame(bilateral_intent_data)
        df_bilateral_intent.to_csv(csv2_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出双边沟通期望投票统计: {csv2_filepath}")
        
        # 3. 期望投票类型汇总统计CSV
        total_support = 0
        total_against = 0
        total_abstain = 0
        total = 0
        
        for key, intents in intent_counts_matrix.items():
            total_support += intents['support']
            total_against += intents['against']
            total_abstain += intents['abstain']
            total += sum(intents.values())
        
        summary_data = [{
            'Intent_Type': 'Support',
            'Intent_Code': 'support',
            'Count': total_support,
            'Percentage': (total_support / total * 100) if total > 0 else 0
        }, {
            'Intent_Type': 'Against',
            'Intent_Code': 'against',
            'Count': total_against,
            'Percentage': (total_against / total * 100) if total > 0 else 0
        }, {
            'Intent_Type': 'Abstain',
            'Intent_Code': 'abstain',
            'Count': total_abstain,
            'Percentage': (total_abstain / total * 100) if total > 0 else 0
        }]
        
        csv3_filename = f"communication_intent_summary_{timestamp}.csv"
        csv3_filepath = data_dir / csv3_filename
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_csv(csv3_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出期望投票类型汇总统计: {csv3_filepath}")
    
    def _export_network_data(self, countries: List[str], pos: Dict, 
                           intent_matrix: Dict, country_intent_stats: Dict,
                           country_comms: List[Dict]):
        """
        导出网络数据到CSV文件
        
        Args:
            countries: 国家列表
            pos: 节点位置字典
            intent_matrix: 倾向性矩阵
            country_intent_stats: 国家期望投票统计
            country_comms: 沟通记录列表
        """
        # 创建数据目录
        data_dir = self.results_dir / "data"
        data_dir.mkdir(exist_ok=True, parents=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. 导出节点信息CSV（国家位置和统计）
        node_data = []
        for i, country in enumerate(countries):
            stats = country_intent_stats[country]
            total = sum(stats.values())
            if total > 0:
                support_pct = stats['support'] / total * 100
                against_pct = stats['against'] / total * 100
                abstain_pct = stats['abstain'] / total * 100
            else:
                support_pct = 0
                against_pct = 0
                abstain_pct = 0
            
            x, y = pos[country]
            node_data.append({
                'Country': country,
                'Position_X': x,
                'Position_Y': y,
                'Total_Outgoing': stats['support'] + stats['against'] + stats['abstain'],
                'Outgoing_Support_Count': stats['support'],
                'Outgoing_Support_Pct': support_pct,
                'Outgoing_Against_Count': stats['against'],
                'Outgoing_Against_Pct': against_pct,
                'Outgoing_Abstain_Count': stats['abstain'],
                'Outgoing_Abstain_Pct': abstain_pct,
                'Network_Centrality_Out': len([c for c in country_comms if c['to'] == country]),
                'Network_Centrality_In': len([c for c in country_comms if c['from'] == country]),
                'Total_Degree': len([c for c in country_comms if c['to'] == country or c['from'] == country])
            })
        
        csv1_filename = f"network_nodes_{timestamp}.csv"
        csv1_filepath = data_dir / csv1_filename
        df_nodes = pd.DataFrame(node_data)
        df_nodes.to_csv(csv1_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出网络节点信息: {csv1_filepath}")
        
        # 2. 导出边信息CSV（沟通记录）
        edge_data = []
        for comm in country_comms:
            from_country = comm["from"]
            to_country = comm["to"]
            
            # 识别沟通意图
            intent = self._detect_communication_intent(comm)
            intent_label = {'support': 'Support', 'against': 'Against', 'abstain': 'Abstain'}.get(intent, 'Unknown')
            intent_code = {'support': 'S', 'against': 'A', 'abstain': 'B'}.get(intent, '?')
            
            # 计算节点位置
            from_pos = pos[from_country]
            to_pos = pos[to_country]
            
            edge_data.append({
                'From_Country': from_country,
                'To_Country': to_country,
                'From_Position_X': from_pos[0],
                'From_Position_Y': from_pos[1],
                'To_Position_X': to_pos[0],
                'To_Position_Y': to_pos[1],
                'Intent_Type': intent_label,
                'Intent_Code': intent_code,
                'Message': comm.get("message", comm.get("content", "")) if isinstance(comm.get("message"), str) else str(comm.get("message", "")),
                'Desired_Vote': comm.get("desired_vote", ""),
                'Timestamp': comm.get("timestamp", "")
            })
        
        csv2_filename = f"network_edges_{timestamp}.csv"
        csv2_filepath = data_dir / csv2_filename
        df_edges = pd.DataFrame(edge_data)
        df_edges.to_csv(csv2_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出网络边信息: {csv2_filepath}")
        
        # 3. 导出网络汇总统计CSV
        total_edges = len(country_comms)
        total_nodes = len(countries)
        
        # 计算网络密度
        max_possible_edges = total_nodes * (total_nodes - 1)
        network_density = total_edges / max_possible_edges if max_possible_edges > 0 else 0
        
        # 计算平均度
        avg_out_degree = sum([len([c for c in country_comms if c['from'] == country]) for country in countries]) / total_nodes
        avg_in_degree = sum([len([c for c in country_comms if c['to'] == country]) for country in countries]) / total_nodes
        
        # 统计边倾向性
        intent_edge_counts = {'support': 0, 'against': 0, 'abstain': 0}
        for comm in country_comms:
            intent = self._detect_communication_intent(comm)
            intent_edge_counts[intent] += 1
        
        summary_stats = [{
            'Metric': 'Total Nodes',
            'Value': total_nodes,
            'Description': 'Number of countries participating in communication'
        }, {
            'Metric': 'Total Edges',
            'Value': total_edges,
            'Description': 'Total number of country-to-country communications'
        }, {
            'Metric': 'Network Density',
            'Value': f"{network_density:.4f}",
            'Description': 'Ratio of actual edges to possible edges'
        }, {
            'Metric': 'Average Out-Degree',
            'Value': f"{avg_out_degree:.2f}",
            'Description': 'Average number of communications sent by each country'
        }, {
            'Metric': 'Average In-Degree',
            'Value': f"{avg_in_degree:.2f}",
            'Description': 'Average number of communications received by each country'
        }, {
            'Metric': 'Support Edges',
            'Value': intent_edge_counts['support'],
            'Percentage': f"{intent_edge_counts['support']/total_edges*100:.1f}%",
            'Description': 'Number of communications with support intent'
        }, {
            'Metric': 'Against Edges',
            'Value': intent_edge_counts['against'],
            'Percentage': f"{intent_edge_counts['against']/total_edges*100:.1f}%",
            'Description': 'Number of communications with against intent'
        }, {
            'Metric': 'Abstain Edges',
            'Value': intent_edge_counts['abstain'],
            'Percentage': f"{intent_edge_counts['abstain']/total_edges*100:.1f}%",
            'Description': 'Number of communications with abstain intent'
        }]
        
        csv3_filename = f"network_summary_{timestamp}.csv"
        csv3_filepath = data_dir / csv3_filename
        df_summary = pd.DataFrame(summary_stats)
        df_summary.to_csv(csv3_filepath, index=False, encoding='utf-8-sig')
        self.logger.info(f"已导出网络汇总统计: {csv3_filepath}")
    
    def _save_analysis_results(self, results: Dict):
        """保存分析结果"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analysis_results_{timestamp}.json"
        filepath = self.results_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"分析结果已保存到: {filepath}")
    
    def create_voting_analysis_charts(self):
        """创建投票分析图表"""
        return self._create_voting_comparison_chart(self.simulation_data)
    
    def create_theory_analysis_charts(self):
        """创建理论分析图表"""
        charts = []
        chart1 = self._create_theory_weights_chart(self.simulation_data)
        chart2 = self._create_theory_scores_heatmap(self.simulation_data)
        if chart1:
            charts.append(chart1)
        if chart2:
            charts.append(chart2)
        return charts
    
    def create_communication_analysis_charts(self):
        """创建沟通分析图表"""
        return self._create_communication_network_chart(self.simulation_data)
    
    def create_accuracy_analysis_charts(self):
        """创建准确率分析图表"""
        return self._create_accuracy_chart(self.simulation_data)
    
    def create_weight_analysis_charts(self):
        """创建权重分析图表"""
        return self._create_weight_evolution_chart(self.simulation_data)
    
    def create_country_comparison_charts(self):
        """创建国家对比图表"""
        return self._create_stance_change_chart(self.simulation_data)
    
    def create_time_series_charts(self):
        """创建时间序列图表"""
        # 可以添加时间序列分析
        return None
    
    def create_comprehensive_dashboard(self):
        """创建综合仪表板"""
        return self._create_dashboard(self.simulation_data)
    
    def generate_detailed_text_report(self):
        """生成详细文本报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"detailed_analysis_report_{timestamp}.md"
        filepath = self.results_dir / filename
        
        # 生成报告内容
        report_content = self._generate_report_content()
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        self.logger.info(f"详细分析报告已保存到: {filepath}")
        return str(filepath)
    
    def _generate_report_content(self) -> str:
        """生成报告内容"""
        content = """# EU Tariff Simulation System Detailed Analysis Report

## Overview

This report provides a comprehensive analysis of the EU electric vehicle tariff voting simulation based on the latest simulation results.

## Key Metrics

"""
        
        # 添加关键指标
        metadata = self.simulation_data.get("simulation_metadata", {})
        analysis = self.simulation_data.get("analysis", {})
        
        content += f"- Simulation Time: {metadata.get('timestamp', 'N/A')}\n"
        content += f"- Participating Countries: {len(metadata.get('countries_participated', []))}\n"
        content += f"- Simulation Duration: {metadata.get('simulation_duration', 'N/A')}\n"
        
        # 投票分析
        voting_analysis = analysis.get("voting_pattern_analysis", {})
        content += f"\n## Voting Analysis\n\n"
        content += f"- Stance Change Rate: {voting_analysis.get('change_rate', 0):.1%}\n"
        content += f"- Stability Index: {voting_analysis.get('stability_index', 0):.2f}\n"
        
        # 沟通分析
        comm_analysis = analysis.get("communication_analysis", {})
        content += f"\n## Communication Analysis\n\n"
        content += f"- Total Communications: {comm_analysis.get('total_communications', 0)}\n"
        content += f"- Country-to-Country Communications: {comm_analysis.get('country_to_country_count', 0)}\n"
        content += f"- EU Commission Communications: {comm_analysis.get('eu_commission_count', 0)}\n"
        
        # 准确率分析
        accuracy_analysis = analysis.get("accuracy_analysis", {})
        content += f"\n## Accuracy Analysis\n\n"
        content += f"- Overall Accuracy: {accuracy_analysis.get('overall_accuracy', 0):.1%}\n"
        content += f"- Countries Evaluated: {accuracy_analysis.get('countries_evaluated', 0)}\n"
        
        # 权重优化
        weight_analysis = analysis.get("weight_optimization_analysis", {})
        content += f"\n## Weight Optimization\n\n"
        content += f"- Countries Optimized: {weight_analysis.get('countries_optimized', 0)}\n"
        content += f"- Average Improvement: {weight_analysis.get('average_improvement', 0):.3f}\n"
        
        content += "\n## Conclusions and Recommendations\n\n"
        content += "Based on the above analysis, it is recommended to further optimize theoretical weights to improve prediction accuracy.\n"
        
        return content

def main():
    """测试函数"""
    # 这里可以添加测试代码
    pass

if __name__ == "__main__":
    main()
